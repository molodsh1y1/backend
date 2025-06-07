import csv
from rest_framework.pagination import PageNumberPagination
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated
from dip.scholar_raw_record.serializers import ScholarRawRecordSerializer
from dip.models import ScholarRawRecord, ScholarAuthor, ScrapingSession
from django.db.models import Q, Count
from django.http import HttpResponse
from datetime import datetime

class ScholarRawRecordPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class ScholarRawRecordViewSet(viewsets.ModelViewSet):
    serializer_class = ScholarRawRecordSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = ScholarRawRecordPagination

    def get_queryset(self):
        queryset = ScholarRawRecord.objects.filter(profile=self.request.profile)

        # Filter by scraping session if provided
        session_id = self.request.query_params.get('session_id')
        if session_id:
            try:
                queryset = queryset.filter(scraping_session_id=int(session_id))
            except (ValueError, TypeError):
                pass

        return queryset.select_related('profile', 'scraping_session').prefetch_related('authors')

    @action(detail=False, methods=['get'], url_path='results/export-csv')
    def export_results_csv(self, request):
        """Export scraping results to CSV for analysts"""
        profile = request.profile

        # Start with base queryset
        queryset = ScholarRawRecord.objects.filter(profile=profile).select_related('profile',
                                                                                   'scraping_session').prefetch_related(
            'authors')

        # Filter by session if provided
        session_id = request.query_params.get('session_id')
        if session_id:
            try:
                queryset = queryset.filter(scraping_session_id=int(session_id))
            except (ValueError, TypeError):
                pass

        # Apply other filters
        query = request.query_params.get('query')
        if query:
            queryset = queryset.filter(
                Q(title__icontains=query) | Q(abstract__icontains=query)
            )

        year_from = request.query_params.get('year_from')
        if year_from:
            try:
                queryset = queryset.filter(publication_year__gte=int(year_from))
            except ValueError:
                pass

        year_to = request.query_params.get('year_to')
        if year_to:
            try:
                queryset = queryset.filter(publication_year__lte=int(year_to))
            except ValueError:
                pass

        venue = request.query_params.get('venue')
        if venue:
            queryset = queryset.filter(venue__icontains=venue)

        min_citations = request.query_params.get('min_citations')
        if min_citations:
            try:
                queryset = queryset.filter(citation_count__gte=int(min_citations))
            except ValueError:
                pass

        # Limit results for performance
        limit = request.query_params.get('limit', 10000)
        try:
            limit = min(int(limit), 50000)  # Max 50k records
        except ValueError:
            limit = 10000

        queryset = queryset.order_by('-scraped_at')[:limit]

        # Create filename with session info
        filename_parts = ["scholar_papers"]
        if session_id:
            filename_parts.append(f"session_{session_id}")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_parts.append(timestamp)
        filename = "_".join(filename_parts) + ".csv"

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        response.write('\ufeff')
        writer = csv.writer(response)

        # Updated headers with session info
        headers = [
            'ID', 'Session ID', 'Session Query', 'Semantic Scholar ID', 'Title', 'Abstract',
            'Publication Year', 'Venue', 'DOI', 'URL', 'PDF URL', 'Citation Count',
            'Reference Count', 'Influential Citation Count', 'Open Access', 'Authors',
            'Author Count', 'Scraped Date', 'Updated Date'
        ]
        writer.writerow(headers)

        # Write data rows with session info
        for paper in queryset:
            authors_list = [author.full_name for author in paper.authors.all()]
            authors_str = "; ".join(authors_list)

            writer.writerow([
                paper.id,
                paper.scraping_session.id if paper.scraping_session else '',
                paper.scraping_session.query if paper.scraping_session else '',
                paper.semantic_scholar_id,
                paper.title,
                paper.abstract,
                paper.publication_year,
                paper.venue,
                paper.doi,
                paper.url,
                paper.pdf_url,
                paper.citation_count,
                paper.reference_count,
                paper.influential_citation_count,
                'Yes' if paper.is_open_access else 'No',
                authors_str,
                len(authors_list),
                paper.scraped_at.strftime('%Y-%m-%d %H:%M:%S'),
                paper.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        return response

    @action(detail=False, methods=['get'], url_path='results/export-excel')
    def export_results_excel(self, request):
        """Export scraping results to Excel for analysts"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return Response({
                "error": "Excel export requires openpyxl library. Please install it."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        from django.http import HttpResponse
        from datetime import datetime
        import io

        profile = request.profile

        # Get filtered queryset with session support
        queryset = ScholarRawRecord.objects.filter(profile=profile).select_related('profile',
                                                                                   'scraping_session').prefetch_related(
            'authors')

        # Filter by session if provided
        session_id = request.query_params.get('session_id')
        if session_id:
            try:
                queryset = queryset.filter(scraping_session_id=int(session_id))
            except (ValueError, TypeError):
                pass

        # Apply other filters
        query = request.query_params.get('query')
        if query:
            queryset = queryset.filter(
                Q(title__icontains=query) | Q(abstract__icontains=query)
            )

        year_from = request.query_params.get('year_from')
        if year_from:
            try:
                queryset = queryset.filter(publication_year__gte=int(year_from))
            except ValueError:
                pass

        year_to = request.query_params.get('year_to')
        if year_to:
            try:
                queryset = queryset.filter(publication_year__lte=int(year_to))
            except ValueError:
                pass

        limit = request.query_params.get('limit', 10000)
        try:
            limit = min(int(limit), 50000)
        except ValueError:
            limit = 10000

        queryset = queryset.order_by('-scraped_at')[:limit]

        # Create Excel workbook
        wb = openpyxl.Workbook()

        # Papers sheet
        ws_papers = wb.active
        ws_papers.title = "Papers"

        # Headers with styling (updated with session info)
        headers = [
            'ID', 'Session ID', 'Session Query', 'Semantic Scholar ID', 'Title', 'Abstract',
            'Publication Year', 'Venue', 'DOI', 'URL', 'PDF URL', 'Citation Count',
            'Reference Count', 'Influential Citation Count', 'Open Access', 'Authors',
            'Author Count', 'Scraped Date', 'Updated Date'
        ]

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws_papers.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data with session info
        for row, paper in enumerate(queryset, 2):
            authors_list = [author.full_name for author in paper.authors.all()]
            authors_str = "; ".join(authors_list)

            # Convert timezone-aware datetimes to naive datetimes for Excel
            scraped_at_naive = paper.scraped_at.replace(tzinfo=None) if paper.scraped_at else None
            updated_at_naive = paper.updated_at.replace(tzinfo=None) if paper.updated_at else None

            data = [
                paper.id,
                paper.scraping_session.id if paper.scraping_session else '',
                paper.scraping_session.query if paper.scraping_session else '',
                paper.semantic_scholar_id,
                paper.title,
                paper.abstract,
                paper.publication_year,
                paper.venue,
                paper.doi,
                paper.url,
                paper.pdf_url,
                paper.citation_count,
                paper.reference_count,
                paper.influential_citation_count,
                'Yes' if paper.is_open_access else 'No',
                authors_str,
                len(authors_list),
                scraped_at_naive,
                updated_at_naive
            ]

            for col, value in enumerate(data, 1):
                ws_papers.cell(row=row, column=col, value=value)

        # Auto-adjust column widths
        for column in ws_papers.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws_papers.column_dimensions[column_letter].width = adjusted_width

        # Authors summary sheet
        ws_authors = wb.create_sheet("Authors Summary")

        # Get author statistics (with session filter if applicable)
        author_stats_query = ScholarAuthor.objects.filter(scholar_raw_records__profile=profile)
        if session_id:
            try:
                author_stats_query = author_stats_query.filter(scholar_raw_records__scraping_session_id=int(session_id))
            except (ValueError, TypeError):
                pass

        author_stats = author_stats_query.annotate(
            papers_count=Count('scholar_raw_records', distinct=True)
        ).order_by('-papers_count')[:100]  # Top 100 authors

        # Authors headers
        author_headers = ['Author Name', 'Papers Count', 'H-Index', 'Total Citations', 'Affiliations']
        for col, header in enumerate(author_headers, 1):
            cell = ws_authors.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Authors data
        for row, author in enumerate(author_stats, 2):
            affiliations_str = "; ".join(author.affiliations) if author.affiliations else ""
            data = [
                author.full_name,
                author.papers_count,
                author.h_index,
                author.citation_count,
                affiliations_str
            ]

            for col, value in enumerate(data, 1):
                ws_authors.cell(row=row, column=col, value=value)

        # Session Summary sheet (if session_id provided)
        if session_id:
            ws_session = wb.create_sheet("Session Summary")

            try:
                from dip.models import ScrapingSession
                session = ScrapingSession.objects.get(id=int(session_id), profile=profile)

                # Session info headers
                session_headers = ['Parameter', 'Value']
                for col, header in enumerate(session_headers, 1):
                    cell = ws_session.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                # Session data
                session_data = [
                    ['Session ID', session.id],
                    ['Query', session.query],
                    ['Year From', session.year_from or 'Any'],
                    ['Year To', session.year_to or 'Any'],
                    ['Limit', session.limit],
                    ['Fields of Study', ', '.join(session.fields_of_study) if session.fields_of_study else 'Any'],
                    ['Publication Types', ', '.join(session.publication_types) if session.publication_types else 'Any'],
                    ['Min Citation Count', session.min_citation_count or 'Any'],
                    ['Open Access Only', 'Yes' if session.open_access_only else 'No'],
                    ['Status', session.status],
                    ['Papers Found', session.papers_found],
                    ['Papers Saved', session.papers_saved],
                    ['Errors Count', session.errors_count],
                    ['Created At', session.created_at.replace(tzinfo=None) if session.created_at else None],
                    ['Completed At', session.completed_at.replace(tzinfo=None) if session.completed_at else None],
                    ['Duration', str(session.duration) if session.duration else 'N/A']
                ]

                for row, (param, value) in enumerate(session_data, 2):
                    ws_session.cell(row=row, column=1, value=param)
                    ws_session.cell(row=row, column=2, value=value)

            except Exception as e:
                # If session not found, just skip this sheet
                wb.remove(ws_session)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Create filename with session info
        filename_parts = ["scholar_analysis"]
        if session_id:
            filename_parts.append(f"session_{session_id}")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_parts.append(timestamp)
        filename = "_".join(filename_parts) + ".xlsx"

        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'], url_path='results/export-authors-csv')
    def export_authors_csv(self, request):
        """Export authors data to CSV for analysts"""
        import csv
        from django.http import HttpResponse
        from datetime import datetime

        profile = request.profile

        # Get authors related to user's papers with session filter
        authors_query = ScholarAuthor.objects.filter(scholar_raw_records__profile=profile)

        session_id = request.query_params.get('session_id')
        if session_id:
            try:
                authors_query = authors_query.filter(scholar_raw_records__scraping_session_id=int(session_id))
            except (ValueError, TypeError):
                pass

        authors = authors_query.distinct().prefetch_related('scholar_raw_records')

        # Create filename with session info
        filename_parts = ["scholar_authors"]
        if session_id:
            filename_parts.append(f"session_{session_id}")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_parts.append(timestamp)
        filename = "_".join(filename_parts) + ".csv"

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        response.write('\ufeff')  # BOM
        writer = csv.writer(response)

        # Headers
        headers = [
            'ID', 'Semantic Scholar ID', 'Full Name', 'URL', 'H-Index',
            'Paper Count', 'Citation Count', 'Affiliations', 'Papers in Dataset',
            'Papers in Session', 'Created Date', 'Updated Date'
        ]
        writer.writerow(headers)

        # Data
        for author in authors:
            papers_in_dataset = author.scholar_raw_records.filter(profile=profile).count()

            # Papers in current session (if session filter applied)
            papers_in_session = papers_in_dataset
            if session_id:
                try:
                    papers_in_session = author.scholar_raw_records.filter(
                        profile=profile,
                        scraping_session_id=int(session_id)
                    ).count()
                except (ValueError, TypeError):
                    papers_in_session = 0

            affiliations_str = "; ".join(author.affiliations) if author.affiliations else ""

            writer.writerow([
                author.id,
                author.semantic_scholar_id,
                author.full_name,
                author.url,
                author.h_index,
                author.paper_count,
                author.citation_count,
                affiliations_str,
                papers_in_dataset,
                papers_in_session,
                author.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                author.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        return response

    @action(detail=False, methods=['get'], url_path='data-ready')
    def check_data_ready(self, request):
        """Check if new data is available for download based on Celery task status"""
        from django.utils import timezone
        from celery.result import AsyncResult

        profile = request.profile

        # Get parameters
        last_check = request.query_params.get('last_check')  # ISO datetime string
        session_id = request.query_params.get('session_id')  # Optional session filter
        task_id = request.query_params.get('task_id')  # Required for task-based logic

        # Start with base queryset
        base_queryset = ScholarRawRecord.objects.filter(profile=profile)

        # Apply session filter if provided
        if session_id:
            try:
                base_queryset = base_queryset.filter(scraping_session_id=int(session_id))
            except (ValueError, TypeError):
                return Response({
                    "error": "Invalid session_id format. Must be integer."
                }, status=status.HTTP_400_BAD_REQUEST)

        # Get Celery task status if task_id provided
        task_status = None
        if task_id:
            try:
                task_result = AsyncResult(task_id)
                task_status = {
                    "task_id": task_id,
                    "status": task_result.status,  # PENDING, STARTED, SUCCESS, FAILURE, RETRY, REVOKED
                    "ready": task_result.ready(),
                    "successful": task_result.successful() if task_result.ready() else None,
                    "result": task_result.result if task_result.ready() else None
                }
            except Exception as e:
                task_status = {
                    "task_id": task_id,
                    "status": "UNKNOWN",
                    "error": str(e)
                }

        if last_check:
            try:
                from django.utils.dateparse import parse_datetime
                last_check_dt = parse_datetime(last_check)

                if not last_check_dt:
                    raise ValueError("Invalid datetime format")

                # Get new papers since last check
                new_papers = base_queryset.filter(scraped_at__gt=last_check_dt).order_by('-scraped_at')
                new_papers_count = new_papers.count()

                # Logic based on Celery task status
                response_data = {
                    "data_ready": new_papers_count > 0,
                    "new_papers_count": new_papers_count,
                    "last_check": last_check,
                    "current_time": timezone.now().isoformat()
                }

                # Add task-specific logic
                if task_status:
                    response_data["task_status"] = task_status

                    # Determine next action based on task status
                    if task_status["status"] == "PENDING":
                        response_data["message"] = "Task is queued, waiting to start"
                        response_data["should_continue_polling"] = True
                        response_data["recommended_interval"] = 5  # seconds

                    elif task_status["status"] == "STARTED":
                        response_data["message"] = "Scraping in progress"
                        response_data["should_continue_polling"] = True
                        response_data["recommended_interval"] = 3  # frequent polling during active scraping

                    elif task_status["status"] == "SUCCESS":
                        response_data["message"] = "Scraping completed successfully"
                        response_data["should_continue_polling"] = False
                        response_data["ready_for_download"] = True

                    elif task_status["status"] == "FAILURE":
                        response_data["message"] = "Scraping failed"
                        response_data["should_continue_polling"] = False
                        response_data["error_details"] = task_status.get("result")

                    elif task_status["status"] == "RETRY":
                        response_data["message"] = "Task is retrying after failure"
                        response_data["should_continue_polling"] = True
                        response_data["recommended_interval"] = 10

                    elif task_status["status"] == "REVOKED":
                        response_data["message"] = "Task was cancelled"
                        response_data["should_continue_polling"] = False

                    else:  # UNKNOWN or other
                        response_data["message"] = "Unknown task status"
                        response_data["should_continue_polling"] = False

                # Add session info if session_id provided
                if session_id:
                    response_data["session_id"] = int(session_id)
                    try:
                        from dip.models import ScrapingSession
                        session = ScrapingSession.objects.get(id=int(session_id), profile=profile)
                        response_data["session_info"] = {
                            "query": session.query,
                            "status": getattr(session, 'status', 'unknown'),
                            "total_papers": getattr(session, 'papers_saved', 0) or 0
                        }
                    except:
                        pass

                # Add preview papers if available
                if new_papers_count > 0:
                    response_data["preview_papers"] = list(new_papers[:3].values(
                        'id', 'title', 'publication_year', 'citation_count', 'scraped_at'
                    ))

                return Response(response_data)

            except Exception as e:
                return Response({
                    "error": "Invalid last_check datetime format. Use ISO format.",
                    "details": str(e)
                }, status=status.HTTP_400_BAD_REQUEST)

        # If no last_check provided, return current state with task status
        total_papers = base_queryset.count()
        recent_threshold = timezone.now() - timezone.timedelta(minutes=10)
        recent_papers = base_queryset.filter(scraped_at__gte=recent_threshold).count()

        response_data = {
            "data_ready": total_papers > 0,
            "total_papers": total_papers,
            "recent_papers": recent_papers,
            "current_time": timezone.now().isoformat(),
            "has_recent_activity": recent_papers > 0
        }

        # Add task status info for initial state
        if task_status:
            response_data["task_status"] = task_status

            if task_status["status"] in ["PENDING", "STARTED"]:
                response_data["should_start_polling"] = True
                response_data["initial_polling_interval"] = 3
            elif task_status["status"] == "SUCCESS":
                response_data["task_completed"] = True
            elif task_status["status"] == "FAILURE":
                response_data["task_failed"] = True
                response_data["error_details"] = task_status.get("result")

        # Add session info if provided
        if session_id:
            response_data["session_id"] = int(session_id)
            try:
                session = ScrapingSession.objects.get(id=int(session_id), profile=profile)
                response_data["session_info"] = {
                    "query": session.query,
                    "status": getattr(session, 'status', 'unknown'),
                    "created_at": session.created_at.isoformat() if session.created_at else None,
                    "completed_at": getattr(session, 'completed_at', None)
                }
                if hasattr(session, 'completed_at') and session.completed_at:
                    response_data["session_info"]["completed_at"] = session.completed_at.isoformat()
            except:
                pass

        return Response(response_data)
